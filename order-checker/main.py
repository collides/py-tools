import os
import openpyxl

warehouse_orders_folder_path = './warehouseOrders'  # 这是包含所有仓库订单数据的文件夹路径
audit_folder_path = './auditFiles'  # 这是包含所有审核文件的文件夹路径

# 参数化变量
order_number_header = '订单号'
supplier_code_header = '经销商代码'
price_header = 'DNP'
quantity_header = '数量'
orig_order_number_header = 'ORIG_ORDNO'
order_no_header = 'ORD_NO'
price_folder_header = 'PRICE'
quantity_folder_header = '数量'

# 硬编码所有字母后面加 2 的前缀替换
order_prefixes_to_replace = {
    'A2': 'A', 'B2': 'B', 'C2': 'C', 'D2': 'D', 'E2': 'E', 'F2': 'F',
    'G2': 'G', 'H2': 'H', 'I2': 'I', 'J2': 'J', 'K2': 'K', 'L2': 'L',
    'M2': 'M', 'N2': 'N', 'O2': 'O', 'P2': 'P', 'Q2': 'Q', 'R2': 'R',
    'S2': 'S', 'T2': 'T', 'U2': 'U', 'V2': 'V', 'W2': 'W', 'X2': 'X',
    'Y2': 'Y', 'Z2': 'Z'
}

header_start_row = 1  # 从第几行开始读取表头
data_start_row = 2  # 从第几行开始读取数据
audit_header_start_row = 4  # 从第几行开始读取 audit 文件的表头
audit_data_start_row = 5  # 从第几行开始读取 audit 文件的数据

def get_all_xlsx_files(folder_path):
    print(f"获取目录中的所有 xlsx 文件: {folder_path}")
    return [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

def replace_prefix(order_number, replacements):
    prefix = order_number[:2]
    if prefix in replacements:
        print(f"替换订单号前缀: {prefix} -> {replacements[prefix]} for {order_number}")
        return replacements[prefix] + order_number[2:]
    return order_number

def aggregate_data_from_folder(folder_path, orig_order_number_header, order_no_header, price_header, quantity_header, header_start_row, data_start_row):
    files = get_all_xlsx_files(folder_path)
    aggregated_data = {}

    for file in files:
        file_path = os.path.join(folder_path, file)
        print(f"处理文件: {file_path}")
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"文件无效或损坏: {file_path}, 错误信息: {e}")
            continue

        for sheet_name in workbook.sheetnames:
            print(f"处理工作表: {sheet_name}")
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=header_start_row, max_row=header_start_row), [])]
            print(f"表头: {headers}")
            if not headers:
                print(f"跳过空表头的工作表: {sheet_name}")
                continue
            for row in sheet.iter_rows(min_row=data_start_row):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                print(f"读取行数据: {row_data}")
                if row_data.get(orig_order_number_header) and row_data.get(order_no_header) and row_data.get(price_header) is not None:
                    key = f"{row_data[orig_order_number_header]}-{row_data[order_no_header][:5]}"
                    if key not in aggregated_data:
                        aggregated_data[key] = {'totalQty': 0, 'totalAmount': 0}
                    aggregated_data[key]['totalQty'] += int(row_data.get(quantity_header, 1))
                    aggregated_data[key]['totalAmount'] += float(row_data[price_header])
                else:
                    print(f"缺少必要字段的数据行: {row_data}")

    print(f"聚合数据完成: {aggregated_data}")
    return aggregated_data

def aggregate_data_from_audit_folder(folder_path, order_number_header, supplier_code_header, price_header, quantity_header, audit_header_start_row, audit_data_start_row):
    files = get_all_xlsx_files(folder_path)
    audit_data = {}

    for file in files:
        file_path = os.path.join(folder_path, file)
        print(f"处理审核文件: {file_path}")
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"审核文件无效或损坏: {file_path}, 错误信息: {e}")
            continue

        file_data = {}
        for sheet_name in workbook.sheetnames:
            print(f"处理审核工作表: {sheet_name}")
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=audit_header_start_row, max_row=audit_header_start_row), [])]
            print(f"审核表头: {headers}")
            if not headers:
                print(f"跳过空表头的审核工作表: {sheet_name}")
                continue
            for row in sheet.iter_rows(min_row=audit_data_start_row):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                print(f"读取审核行数据: {row_data}")
                if row_data.get(order_number_header) and row_data.get(supplier_code_header):
                    order_number = replace_prefix(row_data[order_number_header], order_prefixes_to_replace)
                    key = f"{order_number}-{row_data[supplier_code_header]}"
                    if key not in file_data:
                        file_data[key] = {'totalQty': 0, 'totalAmount': 0}
                    file_data[key]['totalQty'] += int(row_data.get(quantity_header, 1))
                    file_data[key]['totalAmount'] += float(row_data[price_header])
                else:
                    print(f"缺少必要字段的审核数据行: {row_data}")
        
        audit_data[file] = file_data
        print(f"审核数据聚合完成: {file}, 数据: {file_data}")

    return audit_data

def compare_orders(total_orders_data, audit_orders_data):
    all_discrepancies = {}
    all_matched_orders = {}

    print("开始比较数据")
    for file, audit_data in audit_orders_data.items():
        discrepancies = []
        matched_orders = []
        for key in audit_data:
            order_number, service_number = key.split('-')
            audit_total_amount = audit_data[key]['totalAmount']
            audit_total_qty = audit_data[key]['totalQty']

            if key not in total_orders_data:
                discrepancies.append({
                    'orderNumber': order_number,
                    'serviceNumber': service_number,
                    'totalQty': 0,
                    'auditTotalQty': audit_total_qty,
                    'auditTotalAmount': audit_total_amount,
                    'totalAmount': 0
                })
            else:
                total_qty = total_orders_data[key]['totalQty']
                total_amount = total_orders_data[key]['totalAmount']
                if total_amount != audit_total_amount or total_qty != audit_total_qty:
                    discrepancies.append({
                        'orderNumber': order_number,
                        'serviceNumber': service_number,
                        'totalQty': total_qty,
                        'auditTotalQty': audit_total_qty,
                        'auditTotalAmount': audit_total_amount,
                        'totalAmount': total_amount
                    })
                else:
                    matched_orders.append({
                        'orderNumber': order_number,
                        'serviceNumber': service_number,
                        'totalQty': total_qty,
                        'auditTotalQty': audit_total_qty,
                        'auditTotalAmount': audit_total_amount,
                        'totalAmount': total_amount })
        
        all_discrepancies[file] = discrepancies
        all_matched_orders[file] = matched_orders
        print(f"数据比较完成: {file}, 差异: {discrepancies}, 匹配: {matched_orders}")

    return all_discrepancies, all_matched_orders

def generate_result_file(all_discrepancies, all_matched_orders):
    result_file_path = 'comparison_result.txt'
    
    # 如果结果文件已存在，则删除它
    if os.path.exists(result_file_path):
        os.remove(result_file_path)
        print(f"删除已有结果文件: {result_file_path}")

    with open(result_file_path, 'w') as result_file:
        total_discrepancies = sum(len(discrepancies) for discrepancies in all_discrepancies.values())
        total_matched = sum(len(matched_orders) for matched_orders in all_matched_orders.values())
        total_orders = total_discrepancies + total_matched

        result_file.write(f"订单总量: {total_orders}\n\n")

        for file, matched_orders in all_matched_orders.items():
            result_file.write(f"核对没问题的订单 ({file}) {len(matched_orders)}:\n")
            matched_orders.sort(key=lambda x: x['orderNumber'])
            for index, order in enumerate(matched_orders, start=1):
                result_file.write(
                    f"{index}. 订单号: {order['orderNumber']}, 经销商代码: {order['serviceNumber']}, 数量: {order['auditTotalQty']}, 总价: {order['auditTotalAmount']:.2f}\n"
                )

        for file, discrepancies in all_discrepancies.items():
            result_file.write(f"\n核对有问题的订单 ({file}) {len(discrepancies)}:\n")
            discrepancies.sort(key=lambda x: x['orderNumber'])
            for index, discrepancy in enumerate(discrepancies, start=1):
                result_file.write(
                    f"{index}. 订单号: {discrepancy['orderNumber']}, 经销商代码: {discrepancy['serviceNumber']}, 核对文件数量: {discrepancy['auditTotalQty']}, 总订单数量: {discrepancy['totalQty']}, 总订单总价: {discrepancy['totalAmount']:.2f}, 核对文件总价: {discrepancy['auditTotalAmount']:.2f}\n"
                )

    print(f"结果已生成在 {result_file_path}")

# 读取仓库订单文件夹中的所有订单数据并进行数据聚合
print("开始聚合仓库订单文件夹中的订单数据")
total_orders_data = aggregate_data_from_folder(warehouse_orders_folder_path, orig_order_number_header, order_no_header, price_folder_header, quantity_folder_header, header_start_row, data_start_row)

# 读取审核文件夹中的所有订单数据并进行数据聚合
print("开始聚合审核文件夹中的订单数据")
audit_orders_data = aggregate_data_from_audit_folder(audit_folder_path, order_number_header, supplier_code_header, price_header, quantity_header, audit_header_start_row, audit_data_start_row)

# 比较数据
all_discrepancies, all_matched_orders = compare_orders(total_orders_data, audit_orders_data)

# 生成结果文件
generate_result_file(all_discrepancies, all_matched_orders)
